#!/usr/bin/env python3
"""
Excel service for managing the accounting spreadsheet
"""

import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

class ExcelService:
    """
    Service for managing Excel accounting file
    Reads and writes transaction data
    """
    
    def __init__(self, config):
        """Initialize Excel service with configuration"""
        self.config = config
        self.file_path = config.excel_file_path
        self.workbook = None
        self.worksheet = None
        self._load_workbook()
    
    def _load_workbook(self):
        """Load or create Excel workbook"""
        if os.path.exists(self.file_path):
            self.workbook = load_workbook(self.file_path)
            # Use the first sheet
            self.worksheet = self.workbook.active
            print(f"✓ Loaded existing Excel file: {self.file_path}")
        else:
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")
    
    def _get_header_row(self):
        """Find or create header row in worksheet"""
        if self.worksheet.max_row == 0:
            # Create headers
            headers = [
                self.config.excel_date_column,
                self.config.excel_location_column,
                self.config.excel_category_column,
                self.config.excel_description_column,
                self.config.excel_amount_column,
                self.config.excel_receipt_link_column,
                self.config.excel_running_total_column,
            ]
            for col, header in enumerate(headers, 1):
                self.worksheet.cell(row=1, column=col, value=header)
                # Style header
                self.worksheet.cell(row=1, column=col).font = Font(bold=True)
                self.worksheet.cell(row=1, column=col).fill = PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )
            return 1
        return 1
    
    def _find_last_row(self):
        """Find the last row with data"""
        return self.worksheet.max_row
    
    def _get_running_total(self):
        """Get the current running total from last row"""
        last_row = self._find_last_row()
        if last_row <= 1:  # Only header row
            return 0.0
        
        # Get the running total from the last row
        running_total_col = self._get_column_index(self.config.excel_running_total_column)
        last_value = self.worksheet.cell(row=last_row, column=running_total_col).value
        
        try:
            return float(last_value) if last_value else 0.0
        except (ValueError, TypeError):
            return 0.0
    
    def _get_column_index(self, column_name):
        """Get column index by header name"""
        headers = [
            self.config.excel_date_column,
            self.config.excel_location_column,
            self.config.excel_category_column,
            self.config.excel_description_column,
            self.config.excel_amount_column,
            self.config.excel_receipt_link_column,
            self.config.excel_running_total_column,
        ]
        try:
            return headers.index(column_name) + 1
        except ValueError:
            raise ValueError(f"Column '{column_name}' not found in headers")
    
    def add_transactions(self, transactions):
        """
        Add transactions to Excel file
        
        Args:
            transactions (list): List of transaction dictionaries
        
        Returns:
            int: Number of transactions added
        """
        if self.config.dry_run:
            print("[DRY RUN] Would add the following transactions:")
            for txn in transactions:
                print(f"  - {txn['date']}: {txn['merchant']} ${txn['amount']:.2f}")
            return len(transactions)
        
        added_count = 0
        running_total = self._get_running_total()
        
        for txn in transactions:
            last_row = self._find_last_row()
            new_row = last_row + 1
            
            # Calculate new running total
            running_total += txn['amount']
            
            # Insert transaction data
            date_col = self._get_column_index(self.config.excel_date_column)
            location_col = self._get_column_index(self.config.excel_location_column)
            category_col = self._get_column_index(self.config.excel_category_column)
            description_col = self._get_column_index(self.config.excel_description_column)
            amount_col = self._get_column_index(self.config.excel_amount_column)
            receipt_col = self._get_column_index(self.config.excel_receipt_link_column)
            total_col = self._get_column_index(self.config.excel_running_total_column)
            
            self.worksheet.cell(row=new_row, column=date_col, value=txn['date'])
            self.worksheet.cell(row=new_row, column=location_col, value=txn.get('merchant', ''))
            self.worksheet.cell(row=new_row, column=category_col, value=txn.get('category', ''))
            self.worksheet.cell(row=new_row, column=description_col, value=txn.get('description', ''))
            self.worksheet.cell(row=new_row, column=amount_col, value=txn['amount'])
            self.worksheet.cell(row=new_row, column=receipt_col, value=txn.get('receipt_link', ''))
            self.worksheet.cell(row=new_row, column=total_col, value=running_total)
            
            added_count += 1
        
        # Save workbook
        if added_count > 0:
            self.workbook.save(self.file_path)
            print(f"✓ Saved {added_count} transactions to Excel")
        
        return added_count
